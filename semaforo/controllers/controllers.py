# -*- coding: utf-8 -*-
from odoo import http
# import pandas as pd

from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import odoorpc, csv
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


class RegistroObras(http.Controller):
    @http.route('/registro_obras/registro_obras/', auth='public')
    def index(self, **kw):
        estimaciones = http.request.env['control.estimaciones']
        orden = estimaciones.sudo().search([('obra', '=', kw['id'])])

        resultado = []
        obra = ''
        for i in orden:
            resultado.append(i.id)
            obra = i.obra.numero_contrato.contrato
            print(obra)

        # df1 = pd.DataFrame({'ID Estimación': resultado})
        df1 = pd.DataFrame({'ID Estimación': resultado, 'Obra': obra, 'C': 'fg', 'D': 'sdf', 'F': 'fgfg'})

        # Create a Pandas Excel writer using XlsxWriter as the engine.
        writer = pd.ExcelWriter('/tmp/estimaciones.xlsx', engine='xlsxwriter')

        # Write each dataframe to a different worksheet.
        df1.to_excel(writer, sheet_name='Sheet1')

        # Close the Pandas Excel writer and output the Excel file.
        writer.save()

        f = open('/tmp/estimaciones.xlsx', mode="rb")
        return http.request.make_response(f.read(),
                                          [('Content-Type', 'application/octet-stream'),
                                           ('Content-Disposition',
                                            'attachment; filename="{}"'.format('estimaciones.xls'))
                                           ])
        # return "Hello world "+kw['id']


class ReporteAnexop1(http.Controller):
    @http.route('/reporte_anexop1', auth='public')
    def index(self, **kw):
        try:
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo12/extra-addons/finiquito_excel/static/plantilla.xlsx")
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/semaforo/static/anexop1.xlsx")
            sheet = workbook.active
            wb = Workbook()
            # pp.fecha_termino_convenida as "Fecha termino convenida",
            # ej.name as "Ejercicio",
            http.request.env.cr.execute("SELECT pp.id,pp.nombre_contrato,ro.descripcion, " +
                                        "co.name, " +
                                        "pp.total_civa, " +
                                        "pc.fechatermino, " +
                                        "pp.a_fis, " +
                                        "pp.a_fin, " +
                                        "pi.name, " +
                                        "ud.name " +
                                        " FROM partidas_partidas as pp left join proceso_elaboracion_contrato as pc on pp.numero_contrato = pc.id " +
                                        "left join generales_municipios as gm on pp.municipio = gm.id " +
                                        'left join generales_programas_inversion as pi on ' + 'pc."programaInversion"' + ' = pi.id ' +
                                        "left join registro_ejercicio as ej on pc.ejercicio = ej.id " +
                                        "left join contratista_contratista as co on pc.contratista = co.id " +
                                        "left join generales_tipo_obra as too on pc.tipo_obra = too.id " +
                                        "left join registro_programarobra as po on pp.obra = po.id " +
                                        "left join registro_obra as ro on po.obra_planeada = ro.id " +
                                        "left join registro_unidadadminsol as ud on ro.unidadadminsol = ud.id " +
                                        "WHERE ej.name < 2022 ORDER BY pp.id ASC")
            search_obras_ejecucion = http.request.cr.fetchall()

            iterador = 6  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in search_obras_ejecucion:
                iterador += 1
                columna_no = get_column_letter(2)  # COLUMNA CLAVE
                columna_obra = get_column_letter(3)  # COLUMNA CONCEPTO
                columna_proginversion = get_column_letter(4)  # COLUMNA PROGRAMA INVERSION
                columna_unidadadminsol = get_column_letter(5)  # COLUMNA PROGRAMA INVERSION
                columna_contratista = get_column_letter(6)  # COLUMNA UNIDAD
                columna_montociva = get_column_letter(7)  # COLUMNA CANTIDAD
                columna_termino = get_column_letter(8)  # COLUMNA PRECIO UNITARIO
                columna_afis = get_column_letter(9)  # COLUMNA IMPORTE
                columna_afin = get_column_letter(10)  # COLUMNA IMPORTE

                sheet[columna_no + str(iterador)] = i[1]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_obra + str(iterador)] = i[2]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_contratista + str(iterador)] = i[3]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_montociva + str(iterador)] = i[4]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_termino + str(iterador)] = i[5]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_afis + str(iterador)] = i[6]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_afin + str(iterador)] = i[7]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_proginversion + str(iterador)] = i[8]  # AGREGA Programa de inversion
                sheet[columna_unidadadminsol + str(iterador)] = i[9]  # AGREGA UNidad administrativa

                # sheet[columna_importe + str(acum_conceptos)] = i[5]  # AGREGA IMPORTE EN SU POSICION
                # sheet[columna_importe + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            # Save the spreadsheet
            workbook.save("/tmp/anexop1.xlsx")

            # wb.save("/tmp/finiquito.xlsx")

            f = open('/tmp/anexop1.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('anexop1.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class ReporteAnexop2(http.Controller):
    @http.route('/reporte_anexop2', auth='public')
    def index(self, **kw):
        try:
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo12/extra-addons/finiquito_excel/static/plantilla.xlsx")
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/semaforo/static/anexop2.xlsx")
            sheet = workbook.active
            wb = Workbook()
            # pp.fecha_termino_convenida as "Fecha termino convenida",
            # ej.name as "Ejercicio",
            http.request.env.cr.execute("SELECT pp.id,pp.nombre_contrato,ro.descripcion, " +
                                        "co.name, " +
                                        "pp.total_civa, " +
                                        "pc.fechatermino, " +
                                        "pp.fecha1, " +
                                        "pp.a_fin, " +
                                        "pi.name, " +
                                        "ud.name " +
                                        " FROM partidas_partidas as pp left join proceso_elaboracion_contrato as pc on pp.numero_contrato = pc.id " +
                                        "left join generales_municipios as gm on pp.municipio = gm.id " +
                                        'left join generales_programas_inversion as pi on ' + 'pc."programaInversion"' + ' = pi.id ' +
                                        "left join registro_ejercicio as ej on pc.ejercicio = ej.id " +
                                        "left join contratista_contratista as co on pc.contratista = co.id " +
                                        "left join generales_tipo_obra as too on pc.tipo_obra = too.id " +
                                        "left join registro_programarobra as po on pp.obra = po.id " +
                                        "left join registro_obra as ro on po.obra_planeada = ro.id " +
                                        "left join registro_unidadadminsol as ud on ro.unidadadminsol = ud.id " +
                                        "WHERE ej.name < 2022 and pp.fecha9 is not null ORDER BY pp.id ASC")
            search_obras_ejecucion = http.request.cr.fetchall()

            iterador = 6  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in search_obras_ejecucion:
                iterador += 1
                columna_no = get_column_letter(2)  # COLUMNA CLAVE
                columna_obra = get_column_letter(3)  # COLUMNA CONCEPTO
                columna_proginversion = get_column_letter(4)  # COLUMNA PROGRAMA INVERSION
                columna_unidadadminsol = get_column_letter(5)  # COLUMNA PROGRAMA INVERSION
                columna_contratista = get_column_letter(6)  # COLUMNA UNIDAD
                columna_montociva = get_column_letter(7)  # COLUMNA CANTIDAD
                columna_termino = get_column_letter(8)  # COLUMNA PRECIO UNITARIO
                columna_fechaaviso_terminacion = get_column_letter(9)  # COLUMNA IMPORTE
                # columna_afin = get_column_letter(10)  # COLUMNA IMPORTE

                sheet[columna_no + str(iterador)] = i[1]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_obra + str(iterador)] = i[2]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_contratista + str(iterador)] = i[3]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_montociva + str(iterador)] = i[4]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_termino + str(iterador)] = i[5]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_fechaaviso_terminacion + str(iterador)] = i[6]  # AGREGA NOMBRE DEL CONTRATO
                # sheet[columna_afin + str(iterador)] = i[7]  # AGREGA NOMBRE DEL CONTRATO
                sheet[columna_proginversion + str(iterador)] = i[8]  # AGREGA Programa de inversion
                sheet[columna_unidadadminsol + str(iterador)] = i[9]  # AGREGA UNidad administrativa

                # sheet[columna_importe + str(acum_conceptos)] = i[5]  # AGREGA IMPORTE EN SU POSICION
                # sheet[columna_importe + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            # Save the spreadsheet
            workbook.save("/tmp/anexop2.xlsx")

            # wb.save("/tmp/finiquito.xlsx")

            f = open('/tmp/anexop2.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('anexop2.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class ReporteAnexop3(http.Controller):
    @http.route('/reporte_anexop3', auth='public')
    def index(self, **kw):
        try:
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/semaforo/static/anexop3.xlsx")
            sheet = workbook.active
            wb = Workbook()
            http.request.env.cr.execute("SELECT ro.id,ro.numero_obra,ro.descripcion, " +
                                        " ro.localidad, " +
                                        " aoat.estatal, " +
                                        " pi.name, " +
                                        " aoat.federal, " +
                                        " aoat.estatalin, " +
                                        " aoat.federalin, " +
                                        " aoat.municipal, " +
                                        " aoat.municipalin, " +
                                        " aoat.otros, " +
                                        " aoat.otrosin " +
                                        " FROM registro_obra ro " +
                                        "left join registro_ejercicio as ej on ro.ejercicio = ej.id " +
                                        "left join registro_programarobra as po on ro.id = po.obra_planeada " +
                                        'left join generales_programas_inversion as pi on ' + 'po."programaInversion"' + ' = pi.id ' +
                                        "left join autorizacion_obra_anexo_tecnico as aoat on po.id = aoat.concepto " +
                                        "left join registro_unidadadminsol as ud on ro.unidadadminsol = ud.id " +
                                        "WHERE ej.name = 2021 ORDER BY ro.id ASC")
            search_obras_ejecucion = http.request.cr.fetchall()

            iterador = 6  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in search_obras_ejecucion:
                iterador += 1

                buscar_proyecto = http.request.env['registro.proyectoejecutivo'].search_count([('name', '=', i[0])])

                proyecto = ""
                if buscar_proyecto >= 1:
                    proyecto = "Si"
                else:
                    proyecto = "No"

                print(i[4], ' === ', i[7])
                monto_estatal = 0
                monto_estatal_1 = 0
                monto_estatal_2 = 0
                if str(i[4]) == "null":
                    monto_estatal_1 = 0
                elif str(i[7]) == "null":
                    monto_estatal_2 = 0
                elif str(i[4]) != "null":
                    monto_estatal_1 = float(i[4])
                elif str(i[7]) != "null":
                    monto_estatal_2 = float(i[7])
                monto_estatal = monto_estatal_1 + monto_estatal_2

                monto_federal = 0
                monto_federal_1 = 0
                monto_federal_2 = 0
                if str(i[6]) == "null":
                    monto_federal_1 = 0
                elif str(i[8]) == "null":
                    monto_federal_2 = 0
                elif str(i[6]) != "null":
                    monto_federal_1 = float(i[6])
                elif str(i[8]) != "null":
                    monto_federal_2 = float(i[8])
                monto_federal = monto_federal_1 + monto_federal_2

                monto_municipal = 0
                monto_municipal_1 = 0
                monto_municipal_2 = 0
                if str(i[9]) == "null":
                    monto_municipal_1 = 0
                elif str(i[10]) == "null":
                    monto_municipal_2 = 0
                elif str(i[9]) != "null":
                    monto_municipal_1 = float(i[9])
                elif str(i[10]) != "null":
                    monto_municipal_2 = float(i[10])
                monto_municipal = monto_municipal_1 + monto_municipal_2

                monto_otros = 0
                monto_otros_1 = 0
                monto_otros_2 = 0
                if str(i[11]) == "null":
                    monto_otros_1 = 0
                elif str(i[12]) == "null":
                    monto_otros_2 = 0
                elif str(i[11]) != "null":
                    monto_otros_1 = float(i[11])
                elif str(i[12]) != "null":
                    monto_otros_2 = float(i[12])
                monto_otros = monto_otros_1 + monto_otros_2

                columna_no = get_column_letter(2)  # COLUMNA CLAVE
                columna_obra = get_column_letter(3)  # COLUMNA CONCEPTO
                columna_ubicacion = get_column_letter(4)  # COLUMNA PROGRAMA INVERSION
                columna_monto_sonora = get_column_letter(5)  # COLUMNA PROGRAMA INVERSION
                columna_prog_inver = get_column_letter(6)  # COLUMNA PROGRAMA INVERSION
                columna_federal = get_column_letter(7)  # COLUMNA MONTO FEDERAL
                columna_municipal = get_column_letter(8)  # COLUMNA MONTO FEDERAL
                columna_otros = get_column_letter(9)  # COLUMNA MONTO FEDERAL
                columna_proyecto = get_column_letter(10)  # COLUMNA MONTO FEDERAL

                sheet[columna_no + str(iterador)] = i[1]  # AGREGA NUM OBRA
                sheet[columna_obra + str(iterador)] = i[2]  # AGREGA OBRA
                sheet[columna_ubicacion + str(iterador)] = i[3]  # AGREGA UBICACION
                sheet[columna_monto_sonora + str(iterador)] = monto_estatal # float(i[4]) + float(i[7])  # AGREGA MONTO ESTATAL
                sheet[columna_prog_inver + str(iterador)] = i[5]  # AGREGA PROG INVER
                sheet[columna_federal + str(iterador)] = monto_federal # float(i[6]) + float(i[8])  # AGREGA MONTO FEDERAL
                sheet[columna_municipal + str(iterador)] = monto_municipal # float(i[9]) + float(i[10])  # AGREGA MONTO municipal
                sheet[columna_otros + str(iterador)] = monto_otros # float(i[11]) + float(i[12])  # AGREGA MONTO otros
                sheet[columna_proyecto + str(iterador)] = proyecto  # AGREGA SI EXISTE PROYECTO

                # sheet[columna_importe + str(acum_conceptos)] = i[5]  # AGREGA IMPORTE EN SU POSICION
                # sheet[columna_importe + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            # Save the spreadsheet
            workbook.save("/tmp/anexop3.xlsx")

            # wb.save("/tmp/finiquito.xlsx")

            f = open('/tmp/anexop3.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('anexop3.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class ReporteAnexop4(http.Controller):
    @http.route('/reporte_anexop4', auth='public')
    def index(self, **kw):
        try:
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/semaforo/static/anexop4.xlsx")
            sheet = workbook.active
            wb = Workbook()
            http.request.env.cr.execute("SELECT ro.id,ro.numero_obra,ro.descripcion, " +
                                        " ro.localidad, " +
                                        " aoat.estatal, " +
                                        " pi.name, " +
                                        " aoat.federal, " +
                                        " aoat.estatalin, " +
                                        " aoat.federalin, " +
                                        " aoat.municipal, " +
                                        " aoat.municipalin, " +
                                        " aoat.otros, " +
                                        " aoat.otrosin, " +
                                        " of.fecha_de_recibido, " +
                                        " of.fecha_actual, " +
                                        " of.importe " +
                                        " FROM registro_obra ro " +
                                        "left join registro_ejercicio as ej on ro.ejercicio = ej.id " +
                                        "left join registro_programarobra as po on ro.id = po.obra_planeada " +
                                        'left join generales_programas_inversion as pi on ' + 'po."programaInversion"' + ' = pi.id ' +
                                        "left join autorizacion_obra_anexo_tecnico as aoat on po.id = aoat.concepto " +
                                        "left join autorizacion_obra_oficios_de_autorizacion as of on aoat.name = of.id " +
                                        "left join registro_unidadadminsol as ud on ro.unidadadminsol = ud.id " +
                                        "WHERE ej.name = 2022 ORDER BY ro.id ASC")
            search_obras_ejecucion = http.request.cr.fetchall()

            iterador = 6  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in search_obras_ejecucion:
                iterador += 1

                columna_no = get_column_letter(2)  # COLUMNA CLAVE
                columna_obra = get_column_letter(3)  # COLUMNA CONCEPTO
                columna_ubicacion = get_column_letter(4)  # COLUMNA PROGRAMA INVERSION
                columna_monto_sonora = get_column_letter(5)  # COLUMNA PROGRAMA INVERSION
                columna_prog_inver = get_column_letter(6)  # COLUMNA PROGRAMA INVERSION
                columna_federal = get_column_letter(7)  # COLUMNA MONTO FEDERAL
                columna_municipal = get_column_letter(8)  # COLUMNA MONTO FEDERAL
                columna_otros = get_column_letter(9)  # COLUMNA MONTO FEDERAL
                columna_recibido = get_column_letter(10)  # COLUMNA FECHA DE RECIBVIDO
                columna_fecha_autorizacion = get_column_letter(11)  # COLUMNA FECHA DE RECIBVIDO
                columna_monto = get_column_letter(12)  # COLUMNA FECHA DE RECIBVIDO

                sheet[columna_no + str(iterador)] = i[1]  # AGREGA NUM OBRA
                sheet[columna_obra + str(iterador)] = i[2]  # AGREGA OBRA
                sheet[columna_ubicacion + str(iterador)] = i[3]  # AGREGA UBICACION
                sheet[columna_monto_sonora + str(iterador)] = float(i[4]) + float(i[7])  # AGREGA MONTO ESTATAL
                sheet[columna_prog_inver + str(iterador)] = i[5]  # AGREGA PROG INVER
                sheet[columna_federal + str(iterador)] = float(i[6]) + float(i[8])  # AGREGA MONTO FEDERAL
                sheet[columna_municipal + str(iterador)] = float(i[9]) + float(i[10])  # AGREGA MONTO municipal
                sheet[columna_otros + str(iterador)] = float(i[11]) + float(i[12])  # AGREGA MONTO otros
                sheet[columna_recibido + str(iterador)] = i[13]  # AGREGA FECHA RECBIDO
                sheet[columna_fecha_autorizacion + str(iterador)] = i[14]  # AGREGA FECHA AUTORIZACION
                sheet[columna_monto + str(iterador)] = i[15]  # AGREGA FECHA AUTORIZACION

                # sheet[columna_importe + str(acum_conceptos)] = i[5]  # AGREGA IMPORTE EN SU POSICION
                # sheet[columna_importe + str(acum_conceptos)].number_format = '[$$-409]#,##0.00;[RED]-[$$-409]#,##0.00'

            # Save the spreadsheet
            workbook.save("/tmp/anexop4.xlsx")

            # wb.save("/tmp/finiquito.xlsx")

            f = open('/tmp/anexop4.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('anexop4.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)


class ReporteAnexop7(http.Controller):
    @http.route('/reporte_anexop7', auth='public')
    def index(self, **kw):
        try:
            # workbook = load_workbook(filename="/home/gerardo/Developments/odoo12/extra-addons/finiquito_excel/static/plantilla.xlsx")
            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/semaforo/static/anexop7.xlsx")
            sheet = workbook.active
            wb = Workbook()
            http.request.env.cr.execute("SELECT pp.id,pp.nombre_contrato,ro.descripcion, " +
                                        # "co.name, " +
                                        "ro.localidad, " + # [3]
                                        "pi.name, " + # [4]
                                        "ud.name, " + # [5]
                                        "pc.tipo_contrato, " + # [6]
                                        "pc.adjudicacion, " + # [7]
                                        "co.name, " + # [8]
                                        "pc.fecha, " + # [9]
                                        "pc.impcontra, " + # [10]
                                        "pc.fechatermino, " + # [11]
                                        "pp.fecha_anticipos, " + # [12]
                                        "pp.total_anticipo, " + # [13]
                                        "pp.fecha_fianza, " +
                                        "pp.fecha8, " +  # ENTREGA DEL INMUEBLE AL CONTRATISTA
                                        "pp.fecha3, " +  # FECHA DE BITACORA
                                        "pp.a_fis, " +
                                        "pp.a_fin, " +
                                        "pp.fecha_termino_convenida, " +
                                        "pp.fecha1, " +  # FECHA AVISO TERMINACION TRABAJOS
                                        "pp.fecha9, " +  # FECHA inicio finiquito
                                        "pp.fecha10, " +  # FECHA entrega al solicitante
                                        "pc.adjudicacion, " +  # ADJUDICCACION
                                        "co.name " + # CONTRATISTA # [24]

                                        " FROM partidas_partidas as pp left join proceso_elaboracion_contrato as pc on pp.numero_contrato = pc.id " +
                                        "left join generales_municipios as gm on pp.municipio = gm.id " +
                                        'left join generales_programas_inversion as pi on ' + 'pc."programaInversion"' + ' = pi.id ' +
                                        "left join registro_ejercicio as ej on pc.ejercicio = ej.id " +
                                        "left join contratista_contratista as co on pc.contratista = co.id " +
                                        "left join generales_tipo_obra as too on pc.tipo_obra = too.id " +
                                        "left join registro_programarobra as po on pp.obra = po.id " +
                                        "left join registro_obra as ro on po.obra_planeada = ro.id " +
                                        "left join registro_unidadadminsol as ud on ro.unidadadminsol = ud.id " +
                                        "WHERE ej.name = 2022 ORDER BY pp.id ASC")
            search_obras_ejecucion = http.request.cr.fetchall()

            iterador = 6  # INICIA A PARTIR DE LA FILA # 10 Y SE VA ACUMULANDO PARA IR AGREGANDO UN ESPACIO ABAJO CADA OBJETO
            for i in search_obras_ejecucion:
                iterador += 1
                print('xd', i[23])
                fecha_adjudicacion = ""
                if str(i[23]) != "None":
                    adj = http.request.env['proceso.adjudicacion_directa'].browse(int(i[23]))
                    print(adj)
                    for ad in adj:
                        print('si entro', ad.fechaadjudicacion)
                        fecha_adjudicacion = str(ad.fechaadjudicacion)

                print(i[12], ' anticipos ')
                considera_anticipo = ""
                if str(i[12]) == "":
                    considera_anticipo = "No"
                else:
                    considera_anticipo = "Si"

                est = http.request.env['control.estimaciones'].search([('obra', '=', int(i[0]))])
                print('xddd', est)
                acum_estimado = 0
                fecha_pago_anticipo = ""
                for ses in est:
                    best = http.request.env['control.estimaciones'].browse(ses.id)
                    acum_estimado += best.estimado
                    print(i[13], ' aqui ')
                    if float(acum_estimado) >= float(i[13]):
                        fecha_pago_anticipo = best.fecha_inicio_estimacion
                        break
                print('hola')
                pp = http.request.env['partidas.partidas'].search([('id', '=', int(i[0]))])

                residente = ""
                for spp in pp:
                    bspp = http.request.env['partidas.partidas'].browse(spp.id)
                    for r in bspp.residente_obra:
                        residente += r.name + " "
                print(':b')
                num_convenios = http.request.env['proceso.convenios_modificado'].search_count(
                    [('contrato', '=', int(i[0])),('tipo_convenio', '=', 'BOTH')])
                num_est_ordinarias = http.request.env['control.estimaciones'].search_count(
                    [('obra', '=', int(i[0])), ('tipo_estimacion', '=', "1")])
                num_est_extraordinarias = http.request.env['control.estimaciones'].search_count(
                    [('obra', '=', int(i[0])), ('tipo_estimacion', '!=', "1")])

                columna_no = get_column_letter(2)  # COLUMNA Contrato
                columna_obra = get_column_letter(3)  # COLUMNA OBRA
                columna_ubicacion = get_column_letter(4)  # COLUMNA UBICACION
                columna_proginversion = get_column_letter(5)  # COLUMNA PROGRAMA INVERSION
                columna_unidad_solicitante = get_column_letter(6)  # COLUMNA UNIDAD SOLI
                columna_tipo_contrato = get_column_letter(7)  # COLUMNA TIPO CONTRATO
                columna_fecha_adjudicacion = get_column_letter(10)  # COLUMNA FECHA ADJUDICACION
                columna_contratista = get_column_letter(11)  # COLUMNA FECHA ADJUDICACION
                columna_fecha_contrato = get_column_letter(12)  # COLUMNA FECHA CONTRATO
                columna_monto_contrato = get_column_letter(13)  # COLUMNA MONTO CONTRATO
                columna_fechatermino_contrato = get_column_letter(14)  # COLUMNA FECHA TERMINO CONTRATO
                columna_considera_anticipo = get_column_letter(15)  # COLUMNA CONSIDERA ANTICIPO
                columna_fecha_pagoanticipo = get_column_letter(17)  # COLUMNA FECHA PAGO ANTICIPO
                columna_fecha_garantia = get_column_letter(19)  # COLUMNA cumplimiento garantia
                columna_fecha_entregainmueble = get_column_letter(20)  # COLUMNA entrega inmueble
                columna_fecha_bitacora = get_column_letter(21)  # COLUMNA bitacora
                columna_residente = get_column_letter(22)  # COLUMNA residente
                columna_afis = get_column_letter(24)  # COLUMNA avance fisico
                columna_afin = get_column_letter(25)  # COLUMNA avance financiero
                columna_num_convenios = get_column_letter(26)  # COLUMNA NUM CONVENIOS
                columna_vigencia_convenios = get_column_letter(27)  # COLUMNA NUM CONVENIOS
                columna_aviso_terminacion = get_column_letter(28)  # COLUMNA NUM CONVENIOS
                columna_estimaciones_ordinarias = get_column_letter(30)  # COLUMNA NUM CONVENIOS
                columna_estimaciones_extraordinarias = get_column_letter(31)  # COLUMNA NUM CONVENIOS
                columna_fecha_finiquito = get_column_letter(32)  # COLUMNA NUM CONVENIOS
                columna_entrega_solicitante = get_column_letter(33)  # COLUMNA NUM CONVENIOS

                sheet[columna_no + str(iterador)] = i[1]  # AGREGA CONTRATO
                sheet[columna_obra + str(iterador)] = i[2]  # AGREGA OBRA
                sheet[columna_ubicacion + str(iterador)] = i[3]  # AGREGA UBICACION
                sheet[columna_proginversion + str(iterador)] = i[4]  # AGREGA PROGRAMA INVERSION
                sheet[columna_unidad_solicitante + str(iterador)] = i[5]  # AGREGA UNIDAD ADM SOLI
                sheet[columna_tipo_contrato + str(iterador)] = i[6]  # AGREGA TIPO CONTRATO
                sheet[columna_fecha_adjudicacion + str(iterador)] = fecha_adjudicacion  # AGREGA FECHA DE ADJUDICACION
                sheet[columna_contratista + str(iterador)] = i[24]  # AGREGA CONTRATISTA
                sheet[columna_fecha_contrato + str(iterador)] = i[9]  # AGREGA FECHA DE CONTRATO
                sheet[columna_monto_contrato + str(iterador)] = i[10]  # AGREGA MONTO CONTRATO
                sheet[columna_fechatermino_contrato + str(iterador)] = i[11]  # AGREGA FECHA TERMINO CONTRATO
                sheet[columna_considera_anticipo + str(iterador)] = considera_anticipo  # AGREGA CONSIDERA ANTICIPO
                sheet[columna_fecha_pagoanticipo + str(iterador)] = fecha_pago_anticipo  # AGREGA PAGO ANTICIPO
                sheet[columna_fecha_garantia + str(iterador)] = i[14]
                sheet[columna_fecha_entregainmueble + str(iterador)] = i[15]
                sheet[columna_fecha_bitacora + str(iterador)] = i[16]
                sheet[columna_residente + str(iterador)] = residente
                sheet[columna_afis + str(iterador)] = i[17]
                sheet[columna_afin + str(iterador)] = i[18]
                sheet[columna_num_convenios + str(iterador)] = num_convenios
                sheet[columna_vigencia_convenios + str(iterador)] = i[19]
                sheet[columna_aviso_terminacion + str(iterador)] = i[20]
                sheet[columna_estimaciones_ordinarias + str(iterador)] = num_est_ordinarias
                sheet[columna_estimaciones_extraordinarias + str(iterador)] = num_est_extraordinarias
                sheet[columna_fecha_finiquito + str(iterador)] = i[21]
                sheet[columna_entrega_solicitante + str(iterador)] = i[22]

            # Save the spreadsheet
            workbook.save("/tmp/anexop7.xlsx")

            # wb.save("/tmp/finiquito.xlsx")

            f = open('/tmp/anexop7.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('anexop7.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)
