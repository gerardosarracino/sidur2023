# -*- coding: utf-8 -*-
from odoo import http

from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.styles import colors
from openpyxl.styles import Font, Color, NamedStyle
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


class Memorandum(http.Controller):
    @http.route('/memorandum/memorandum/', auth='public')
    def index(self, id):
        try:
            memo = http.request.env['expediente.memorandum'].sudo().search([('id', '=', int(id))])
            thin = Side(border_style="thin", color="000000")
            fecha_memo = ''
            no_memo = ''

            workbook = load_workbook(filename="/usr/lib/python3/dist-packages/odoo/odoo-extra-addons/control_expediente/static/memo_plantilla.xlsx")
            sheet = workbook.active
            wb = Workbook()

            for i in memo:
                b_memo = http.request.env['expediente.memorandum'].browse(i.id)

                cont = 0
                for x in b_memo.tabla_reporte:
                    cont += 1
                    columna_id = get_column_letter(1)
                    columna_contrato = get_column_letter(2)
                    columna_documento = get_column_letter(3)
                    columna_documento2 = get_column_letter(4)
                    columna_observaciones = get_column_letter(5)
                    columna_observaciones2 = get_column_letter(6)

                    sheet[columna_id + str(cont + 8)] = str(x.id_numeracion)
                    sheet[columna_id + str(cont + 8)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    sheet[columna_contrato + str(cont + 8)] = str(x.contrato_id.contrato)
                    sheet[columna_contrato + str(cont + 8)].border = Border(top=thin, left=thin, right=thin,
                                                                            bottom=thin)
                    sheet[columna_documento + str(cont + 8)] = str(x.nombre_documento.nombre_documento)
                    sheet.merge_cells("" + columna_documento + str(cont + 8) + ':' + columna_documento2 + str(cont + 8))
                    sheet[columna_documento + str(cont + 8)].border = Border(top=thin, left=thin, right=thin,
                                                                             bottom=thin)
                    sheet[columna_observaciones + str(cont + 8)] = str(x.observaciones)
                    sheet.merge_cells(
                        "" + columna_observaciones + str(cont + 8) + ':' + columna_observaciones2 + str(cont + 8))
                    sheet[columna_observaciones + str(cont + 8)].border = Border(top=thin, left=thin, right=thin,
                                                                                 bottom=thin)

                texto = 'RECIBE EL (LA) RESPONSABLE DEL ARCHIVO DGEO PARA SU REVISIÓN DETALLADA'
                texto2 = 'ESTE ACUSE NO ES VÁLIDO \n SI NO TIENE EL SELLO DEL ARCHIVO.'
                fecha_memo = b_memo.fecha_memo
                no_memo = b_memo.id_reporte
                columna_texto = get_column_letter(1)
                columna_texto2 = get_column_letter(6)
                sheet[columna_texto + str(cont + 10)] = texto
                sheet[columna_texto + str(cont + 10)].alignment = Alignment(horizontal="center", vertical="center")
                sheet[columna_texto + str(cont + 19)] = texto2
                sheet[columna_texto + str(cont + 19)].alignment = Alignment(horizontal="center", vertical="center")
                columna_fecha = get_column_letter(6)
                columna_no = get_column_letter(6)
                sheet.merge_cells("" + columna_texto + str(cont + 10) + ':' + columna_texto2 + str(cont + 10))
                sheet.merge_cells("" + columna_texto + str(cont + 19) + ':' + columna_texto2 + str(cont + 19))
                sheet[columna_fecha + str(5)] = fecha_memo
                sheet[columna_no + str(cont + 3)] = no_memo

            # Save the spreadsheet
            workbook.save("/tmp/reporte_memorandum.xlsx")

            # prs.save('/tmp/test.pptx')
            f = open('/tmp/reporte_memorandum.xlsx', mode="rb")
            return http.request.make_response(f.read(),
                                              [('Content-Type', 'application/octet-stream'),
                                               ('Content-Disposition',
                                                'attachment; filename="{}"'.format('reporte_memorandum.xlsx'))
                                               ])

        except Exception as e:
            return "Upss! algo salio mal en: " + str(e)
